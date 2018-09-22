# -*- coding: utf-8 -*-
"""
Created on Tue Jul 24 09:29:59 2018

@author: Nigel
"""

import os
import pandas
import numpy as np
from CitiTech_BuyerClass_final import *

# Check current directory
currentDir = os.getcwd()

# Enter directory to save excel files in
#path = 'C:\\Users\\Nigel\\Documents\\GitHub\\citi_tech_project\\exampleApp'
os.chdir(currentDir)

writer = pandas.ExcelWriter('testA.xlsx',  engine='openpyxl')

# Import variables from CitiTech_BuyerClass_Final.py
age = CitiTech_BuyerClass_final.buyer.age
gender = CitiTech_BuyerClass_final.buyer.gender
monthlyIncome = CitiTech_BuyerClass_final.buyer.monthlyIncome
expenses = CitiTech_BuyerClass_final.buyer.expenses
personalSavings = CitiTech_BuyerClass_final.buyer.personalSavings
availDownPay = CitiTech_BuyerClass_final.buyer.availDownPay
existingLoan = CitiTech_BuyerClass_final.buyer.existingLoan
propertyLocation = CitiTech_BuyerClass_final.buyer.propertyLocation

# Process datetime data
datetime = CitiTech_BuyerClass_final.buyer.datetime
datetime = datetime.timetuple()
time = []
for x in datetime:
    time.append(str(x))
timeStr = '-'.join(time)

# Summary of user data in list format
userInfo = [timeStr, age, gender, monthlyIncome, expenses, personalSavings, availDownPay, existingLoan, propertyLocation]

# Converts to dataframe
df = pandas.DataFrame()
dfline = pandas.DataFrame(np.reshape(userInfo,(1,9)))
df = df.append(dfline)
print(df)

# Checks if 'testA.xlsx' is in the directory, if yes then appends, if no then create new file
fileList = os.listdir()
if 'testA.xlsx' in fileList:
    from openpyxl import load_workbook
    wb = load_workbook('testA.xlsx')
    #type(wb)
    ws = wb.active
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    print(row_count)
else:
    row_count = 0

# Appends/writes to excel
df.to_excel(writer, header=False, index=False, startrow = row_count)
writer.save()
